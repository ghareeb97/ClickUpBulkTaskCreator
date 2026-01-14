"""
Excel parser for HNDL User Stories format.
Handles multi-sheet workbooks with Epics and User Story sheets.
"""

import re
import io
from openpyxl import load_workbook
from typing import Dict, List, Tuple


def get_sheet_names(workbook_bytes: bytes) -> List[str]:
    """Load workbook from bytes and return list of sheet names."""
    wb = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)
    return wb.sheetnames


def parse_linked_user_stories(linked_str: str) -> List[str]:
    """
    Parse the 'Linked User Stories' column which can contain:
    - Comma-separated IDs: "US-Login-1, US-Login-2, US-Login-3"
    - Range format: "US-VR-15 -> US-VR-17" (expands to US-VR-15, US-VR-16, US-VR-17)
    - Mixed: "US-Login-1, US-VR-15 -> US-VR-17, US-Auth-5"

    Returns list of individual User Story IDs.
    """
    if not linked_str or str(linked_str).strip() == "" or linked_str is None:
        return []

    result = []
    linked_str = str(linked_str).strip()

    # Split by comma first
    parts = [p.strip() for p in linked_str.split(",")]

    for part in parts:
        if not part:
            continue

        # Check if this is a range (contains -> or arrow variants)
        range_match = re.match(
            r"(US-([A-Za-z]+)-(\d+))\s*(?:->|→|to)\s*(US-([A-Za-z]+)-(\d+))",
            part, re.IGNORECASE
        )

        if range_match:
            # Extract prefix and numeric range
            start_prefix = range_match.group(2)
            start_num = int(range_match.group(3))
            end_prefix = range_match.group(5)
            end_num = int(range_match.group(6))

            # Ensure prefixes match (they should for valid ranges)
            if start_prefix.upper() == end_prefix.upper():
                prefix = f"US-{start_prefix}-"
                # Generate all IDs in range
                for num in range(start_num, end_num + 1):
                    result.append(f"{prefix}{num}")
            else:
                # Prefixes don't match - treat as single item
                result.append(part)
        else:
            # Single ID - validate format and add
            if re.match(r"US-[A-Za-z]+-\d+", part, re.IGNORECASE):
                result.append(part)

    return result


def parse_epics_sheet(wb, sheet_name: str = "Epics") -> Dict[str, dict]:
    """
    Parse the Epics sheet and build a mapping from User Story ID to Epic info.

    Expected columns: Epic ID, Epic Title, Epic Description, Linked User Stories

    Returns: Dict mapping User Story ID -> Epic info dict with:
        - epic_id: "EPIC-Login-1"
        - epic_title: "Entry Points & Navigation"
        - epic_description: "Allow users to reach..."
        - epic_label: "EPIC-Login-1: Entry Points & Navigation" (for dropdown)
    """
    epic_mapping = {}

    if sheet_name not in wb.sheetnames:
        return epic_mapping

    ws = wb[sheet_name]

    # Find column indices from header row
    headers = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[str(cell.value).strip().lower()] = idx

    # Required columns
    id_col = headers.get("epic id")
    title_col = headers.get("epic title")
    desc_col = headers.get("epic description")
    linked_col = headers.get("linked user stories")

    if not title_col or not linked_col:
        raise ValueError(
            f"Epics sheet missing required columns. "
            f"Found: {list(headers.keys())}. "
            f"Required: 'Epic Title', 'Linked User Stories'"
        )

    # Parse each row (skip header)
    for row in ws.iter_rows(min_row=2):
        epic_id = row[id_col - 1].value if id_col else None
        epic_title = row[title_col - 1].value
        epic_desc = row[desc_col - 1].value if desc_col else None
        linked_stories = row[linked_col - 1].value

        if not epic_title:
            continue

        epic_id = str(epic_id).strip() if epic_id else ""
        epic_title = str(epic_title).strip()
        epic_desc = str(epic_desc).strip() if epic_desc else ""

        # Build Epic label for dropdown: "EPIC-Login-1: Entry Points & Navigation"
        if epic_id:
            epic_label = f"{epic_id}: {epic_title}"
        else:
            epic_label = epic_title

        epic_info = {
            "epic_id": epic_id,
            "epic_title": epic_title,
            "epic_description": epic_desc,
            "epic_label": epic_label
        }

        # Parse linked user stories and map each to this epic
        user_story_ids = parse_linked_user_stories(linked_stories)
        for us_id in user_story_ids:
            epic_mapping[us_id.upper()] = epic_info

    return epic_mapping


def parse_user_story_sheet(wb, sheet_name: str, epic_mapping: Dict[str, dict]) -> List[dict]:
    """
    Parse a single User Story sheet and return list of task dictionaries.

    Expected columns: User Story ID, User Story Title, User Story,
                     Acceptance Criteria, Revised?, Status

    Returns list of dicts with:
    - name: "US-Login-1: Landing Page" (ID + Title)
    - description: User Story text + Acceptance Criteria + Epic Info
    - epic: Epic label for dropdown (e.g., "EPIC-Login-1: Entry Points & Navigation")
    - epic_info: Full epic info dict
    - user_story_id: Original ID for reference
    """
    tasks = []

    if sheet_name not in wb.sheetnames:
        return tasks

    ws = wb[sheet_name]

    # Find column indices from header row
    headers = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value:
            headers[str(cell.value).strip().lower()] = idx

    # Map common column name variations
    id_col = (headers.get("user story id") or
              headers.get("us id") or
              headers.get("id"))
    title_col = (headers.get("user story title") or
                 headers.get("title"))
    story_col = (headers.get("user story") or
                 headers.get("story") or
                 headers.get("description"))
    ac_col = (headers.get("acceptance criteria") or
              headers.get("ac"))

    if not id_col or not title_col:
        raise ValueError(
            f"Sheet '{sheet_name}' missing required columns. "
            f"Found: {list(headers.keys())}. "
            f"Required: 'User Story ID', 'User Story Title'"
        )

    # Parse each row (skip header)
    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        us_id = row[id_col - 1].value
        us_title = row[title_col - 1].value

        # Skip empty rows
        if not us_id or not us_title:
            continue

        us_id = str(us_id).strip()
        us_title = str(us_title).strip()

        # Build task name: "US-Login-1: Landing Page"
        task_name = f"{us_id}: {us_title}"

        # Look up Epic from mapping (case-insensitive)
        epic_info = epic_mapping.get(us_id.upper())

        # Build description
        description_parts = []

        if story_col and row[story_col - 1].value:
            description_parts.append(str(row[story_col - 1].value).strip())

        if ac_col and row[ac_col - 1].value:
            ac_text = str(row[ac_col - 1].value).strip()
            if ac_text:
                description_parts.append(f"\n\n## Acceptance Criteria\n{ac_text}")

        # Add Epic info section to description
        if epic_info:
            epic_section = f"\n\n---\n## Epic Info\n**{epic_info['epic_label']}**"
            if epic_info.get('epic_description'):
                epic_section += f"\n{epic_info['epic_description']}"
            description_parts.append(epic_section)

        description = "".join(description_parts)

        tasks.append({
            "name": task_name,
            "description": description,
            "epic": epic_info["epic_label"] if epic_info else None,
            "epic_info": epic_info,
            "user_story_id": us_id,
            "_source_sheet": sheet_name,
            "_source_row": row_num
        })

    return tasks


def parse_excel_workbook(
    workbook_bytes: bytes,
    selected_sheets: List[str]
) -> Tuple[List[dict], Dict[str, int]]:
    """
    Main entry point for Excel parsing.

    Args:
        workbook_bytes: Raw bytes of uploaded Excel file
        selected_sheets: List of sheet names to import (User Story sheets only)

    Returns:
        Tuple of:
        - List of task dictionaries ready for ClickUp API
        - Dict of stats: {"total_tasks": N, "with_epic": M, "sheets_processed": K}
    """
    wb = load_workbook(io.BytesIO(workbook_bytes), read_only=True, data_only=True)

    # First, parse Epics sheet to build the mapping
    epic_mapping = parse_epics_sheet(wb)

    # Parse selected User Story sheets
    all_tasks = []
    for sheet_name in selected_sheets:
        sheet_tasks = parse_user_story_sheet(wb, sheet_name, epic_mapping)
        all_tasks.extend(sheet_tasks)

    # Calculate stats
    unique_epics = set(info["epic_label"] for info in epic_mapping.values())
    stats = {
        "total_tasks": len(all_tasks),
        "with_epic": sum(1 for t in all_tasks if t.get("epic")),
        "sheets_processed": len(selected_sheets),
        "total_epics": len(unique_epics)
    }

    return all_tasks, stats
